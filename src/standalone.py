"""
CryptoHunter - Standalone Server
Full pipeline with XGBoost filter, GNN inference, and Angr verification.
"""

import os
import uuid
import json
import hashlib
import asyncio
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Set

from fastapi import FastAPI, File, UploadFile, HTTPException, WebSocket, WebSocketDisconnect
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
import aiofiles

# WebSocket connection manager for real-time progress
class ConnectionManager:
    def __init__(self):
        self.active_connections: Dict[str, List[WebSocket]] = {}
    
    async def connect(self, websocket: WebSocket, job_id: str):
        await websocket.accept()
        if job_id not in self.active_connections:
            self.active_connections[job_id] = []
        self.active_connections[job_id].append(websocket)
    
    def disconnect(self, websocket: WebSocket, job_id: str):
        if job_id in self.active_connections:
            self.active_connections[job_id].remove(websocket)
    
    async def send_progress(self, job_id: str, message: dict):
        if job_id in self.active_connections:
            for ws in self.active_connections[job_id]:
                try:
                    await ws.send_json(message)
                except:
                    pass

manager = ConnectionManager()


# Paths
BASE_DIR = Path(__file__).parent
FRONTEND_DIR = BASE_DIR / "frontend" / "dist"
UPLOAD_DIR = BASE_DIR / "uploads"
RESULTS_DIR = BASE_DIR / "results"

# Create dirs
UPLOAD_DIR.mkdir(exist_ok=True)
RESULTS_DIR.mkdir(exist_ok=True)

# Create app
app = FastAPI(
    title="CryptoHunter",
    description="AI/ML Firmware Crypto Detection",
    version="2.0.0"
)

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Serve static assets if frontend is built
if (FRONTEND_DIR / "assets").exists():
    app.mount("/assets", StaticFiles(directory=FRONTEND_DIR / "assets"), name="assets")


@app.get("/")
async def root():
    """Serve frontend or API info."""
    index_path = FRONTEND_DIR / "index.html"
    if index_path.exists():
        return FileResponse(index_path)
    
    return HTMLResponse("""
    <!DOCTYPE html>
    <html>
    <head>
        <title>CryptoHunter</title>
        <style>
            * { box-sizing: border-box; }
            body { font-family: 'Segoe UI', Arial, sans-serif; max-width: 900px; margin: 0 auto; padding: 20px; background: #1a1a2e; color: #eee; }
            h1 { color: #00d9ff; margin-bottom: 5px; }
            .subtitle { color: #888; margin-bottom: 30px; }
            .upload-box { border: 2px dashed #00d9ff; padding: 40px; text-align: center; border-radius: 15px; margin: 20px 0; background: #16213e; transition: all 0.3s; }
            .upload-box:hover { border-color: #00ff88; background: #1a2744; }
            .upload-box.dragover { border-color: #00ff88; background: #1a3344; }
            input[type="file"] { font-size: 16px; color: #eee; }
            .btn { background: linear-gradient(135deg, #00d9ff, #00ff88); color: #000; padding: 12px 30px; border: none; border-radius: 8px; cursor: pointer; font-size: 16px; font-weight: bold; }
            .btn:hover { transform: scale(1.05); }
            .btn:disabled { opacity: 0.5; cursor: not-allowed; }
            #status { color: #00ff88; font-size: 18px; margin-top: 15px; }
            .results-container { background: #16213e; border-radius: 15px; padding: 25px; margin-top: 20px; }
            .section { margin-bottom: 25px; }
            .section h3 { color: #00d9ff; border-bottom: 1px solid #333; padding-bottom: 10px; margin-bottom: 15px; }
            .crypto-item { background: #0f3460; padding: 15px; border-radius: 10px; margin-bottom: 10px; display: flex; justify-content: space-between; align-items: center; }
            .crypto-name { font-weight: bold; color: #fff; }
            .crypto-type { color: #00d9ff; font-size: 14px; }
            .crypto-indicator { color: #888; font-size: 12px; margin-top: 5px; }
            .confidence { background: #00ff88; color: #000; padding: 5px 12px; border-radius: 20px; font-weight: bold; }
            .confidence.low { background: #ff6b6b; }
            .confidence.medium { background: #ffd93d; }
            .protocol-item { background: #1a3a5c; padding: 15px; border-radius: 10px; margin-bottom: 10px; }
            .protocol-name { font-weight: bold; color: #ffd93d; font-size: 18px; }
            .protocol-desc { color: #aaa; }
            .summary-grid { display: grid; grid-template-columns: repeat(3, 1fr); gap: 15px; }
            .summary-card { background: #0f3460; padding: 20px; border-radius: 10px; text-align: center; }
            .summary-value { font-size: 32px; font-weight: bold; color: #00d9ff; }
            .summary-label { color: #888; font-size: 14px; }
            .no-results { color: #888; text-align: center; padding: 40px; }
            pre { background: #0a0a15; padding: 15px; border-radius: 10px; overflow-x: auto; font-size: 12px; max-height: 300px; }
        </style>
    </head>
    <body>
        <h1> CryptoHunter</h1>
        <p class="subtitle">AI-Powered Cryptographic Primitive Detection in Firmware</p>
        
        <div class="upload-box" id="dropzone">
            <h3> Upload Firmware</h3>
            <p>Drag & drop a firmware file here, or click to select</p>
            <form id="uploadForm" enctype="multipart/form-data">
                <input type="file" name="file" id="fileInput" accept=".bin,.elf,.o,.so,.exe,.img,.fw">
                <br><br>
                <button type="submit" class="btn" id="submitBtn"> Analyze Firmware</button>
            </form>
            <p id="status"></p>
        </div>
        
        <div id="results"></div>
        
        <script>
            const dropzone = document.getElementById('dropzone');
            const fileInput = document.getElementById('fileInput');
            
            dropzone.ondragover = (e) => { e.preventDefault(); dropzone.classList.add('dragover'); };
            dropzone.ondragleave = () => dropzone.classList.remove('dragover');
            dropzone.ondrop = (e) => {
                e.preventDefault();
                dropzone.classList.remove('dragover');
                fileInput.files = e.dataTransfer.files;
                document.getElementById('uploadForm').dispatchEvent(new Event('submit'));
            };
            
            document.getElementById('uploadForm').onsubmit = async (e) => {
                e.preventDefault();
                const file = fileInput.files[0];
                if (!file) { alert('Please select a file'); return; }
                
                const formData = new FormData();
                formData.append('file', file);
                
                document.getElementById('status').innerText = '⏳ Analyzing ' + file.name + '...';
                document.getElementById('submitBtn').disabled = true;
                
                try {
                    const res = await fetch('/api/analyze', { method: 'POST', body: formData });
                    const data = await res.json();
                    document.getElementById('status').innerText = ' Analysis Complete!';
                    displayResults(data);
                } catch (err) {
                    document.getElementById('status').innerText = ' Error: ' + err.message;
                }
                document.getElementById('submitBtn').disabled = false;
            };
            
            function displayResults(data) {
                const cryptoCount = data.summary?.crypto_count || 0;
                const arch = data.summary?.architecture || 'Unknown';
                const protocols = data.protocols || [];
                const classifications = data.classifications || [];
                
                let html = '<div class="results-container">';
                
                // Summary
                html += '<div class="section"><h3> Summary</h3><div class="summary-grid">';
                html += '<div class="summary-card"><div class="summary-value">' + cryptoCount + '</div><div class="summary-label">Crypto Primitives</div></div>';
                html += '<div class="summary-card"><div class="summary-value">' + protocols.length + '</div><div class="summary-label">Protocols</div></div>';
                html += '<div class="summary-card"><div class="summary-value">' + arch + '</div><div class="summary-label">Architecture</div></div>';
                html += '</div></div>';
                
                // Classifications
                html += '<div class="section"><h3> Detected Crypto Primitives</h3>';
                if (classifications.length > 0) {
                    classifications.forEach(c => {
                        const confClass = c.confidence >= 0.85 ? '' : c.confidence >= 0.70 ? 'medium' : 'low';
                        html += '<div class="crypto-item">';
                        html += '<div><div class="crypto-name">' + c.class_name + '</div><div class="crypto-indicator">' + (c.indicator || c.name) + '</div></div>';
                        html += '<div class="confidence ' + confClass + '">' + Math.round(c.confidence * 100) + '%</div>';
                        html += '</div>';
                    });
                } else {
                    html += '<div class="no-results">No crypto primitives detected</div>';
                }
                html += '</div>';
                
                // Protocols
                if (protocols.length > 0) {
                    html += '<div class="section"><h3> Detected Protocols</h3>';
                    protocols.forEach(p => {
                        html += '<div class="protocol-item">';
                        html += '<div class="protocol-name">' + p.name + '</div>';
                        html += '<div class="protocol-desc">' + (p.description || '') + ' (' + Math.round(p.confidence * 100) + '% confidence)</div>';
                        html += '</div>';
                    });
                    html += '</div>';
                }
                
                // Raw JSON
                html += '<div class="section"><h3> Raw JSON</h3><pre>' + JSON.stringify(data, null, 2) + '</pre></div>';
                
                html += '</div>';
                document.getElementById('results').innerHTML = html;
            }
        </script>
    </body>
    </html>
    """)


@app.get("/api/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}


@app.post("/api/analyze")
async def analyze(file: UploadFile = File(...)):
    """Upload and analyze firmware."""
    job_id = str(uuid.uuid4())[:8]
    
    # Save file
    file_path = UPLOAD_DIR / f"{job_id}_{file.filename}"
    async with aiofiles.open(file_path, 'wb') as f:
        content = await file.read()
        await f.write(content)
    
    file_hash = hashlib.sha256(content).hexdigest()
    
    # Run REAL analysis with GNN model
    result = await run_real_analysis(job_id, str(file_path), file.filename, len(content), file_hash)
    
    # Save result
    with open(RESULTS_DIR / f"{job_id}.json", 'w') as f:
        json.dump(result, f, indent=2)
    
    return result


async def run_real_analysis(job_id, file_path, filename, size, file_hash):
    """Run actual GNN-based crypto detection with Ghidra + auto-extraction."""
    import sys
    import subprocess
    import shutil
    import tempfile
    import lzma
    import gzip
    import zipfile
    
    # Add parent directory to path for imports
    parent_dir = str(BASE_DIR.parent)
    if parent_dir not in sys.path:
        sys.path.insert(0, parent_dir)
    
    results = {
        "job_id": job_id,
        "status": "completed",
        "filename": filename,
        "size": size,
        "sha256": file_hash,
        "timestamp": datetime.utcnow().isoformat(),
        "classifications": [],
        "protocols": [],
        "summary": {},
        "extraction": None,
        "ghidra_analysis": None  # Will contain Ghidra results if run
    }
    
    # ==================================================================
    # FULL GHIDRA PIPELINE - For proper crypto detection
    # ==================================================================
    ghidra_path = os.environ.get('GHIDRA_PATH', r'D:\ghidra_11.4.2_PUBLIC')
    ghidra_available = os.path.exists(ghidra_path)
    
    if ghidra_available:
        try:
            # Import the full analysis pipeline
            from run_full_analysis import analyze
            
            # Configure output directory
            output_dir = RESULTS_DIR / job_id
            output_dir.mkdir(exist_ok=True)
            
            # Run full Ghidra + GNN analysis
            print(f" Running full Ghidra analysis for {filename}...")
            
            # Output path for the JSON report
            output_json = str(output_dir / f"{job_id}_report.json")
            
            ghidra_result = analyze(
                binary_path=file_path,
                output_path=output_json,
                is_firmware=True  # Enable firmware extraction
            )
            
            if ghidra_result:
                # Map Ghidra results to our response format
                metadata = ghidra_result.get("metadata", {})
                
                results["ghidra_analysis"] = {
                    "performed": True,
                    "functions_analyzed": metadata.get("total_functions", 0),
                    "crypto_functions": metadata.get("crypto_functions", 0)
                }
                
                # Use Ghidra classifications directly
                ghidra_classifications = ghidra_result.get("classifications", [])
                for func in ghidra_classifications:
                    if func.get("class_id", 0) != 0:  # Non-crypto functions excluded
                        results["classifications"].append({
                            "name": func.get("name", "unknown"),
                            "class_id": func.get("class_id", 0),
                            "class_name": func.get("class_name", "Unknown"),
                            "confidence": func.get("confidence", 0),
                            "indicator": f"GNN @ {func.get('entry', 'N/A')}"
                        })
                
                # Use Ghidra protocols
                if "protocols" in ghidra_result:
                    results["protocols"] = ghidra_result["protocols"]
                
                # Summary from Ghidra
                ghidra_summary = ghidra_result.get("summary", {})
                
                # Also run our dual-method architecture detection for enriched results
                arch_detection_details = detect_architecture_detailed(file_path)
                if arch_detection_details and arch_detection_details.get('final', {}).get('architecture') != 'Unknown':
                    results["architecture_detection"] = {
                        "method_1": {
                            "name": arch_detection_details['method_1'].get('name', 'Capstone Disassembly'),
                            "architecture": arch_detection_details['method_1'].get('architecture'),
                            "confidence": arch_detection_details['method_1'].get('confidence', 0)
                        },
                        "method_2": {
                            "name": arch_detection_details['method_2'].get('name', 'Header + Prologue'),
                            "architecture": arch_detection_details['method_2'].get('architecture'),
                            "confidence": arch_detection_details['method_2'].get('confidence', 0)
                        },
                        "final": arch_detection_details.get('final', {})
                    }
                    # Use our detection if Ghidra didn't provide architecture
                    detected_arch = metadata.get("binary", "Unknown")
                    if detected_arch == "Unknown":
                        final = arch_detection_details.get('final', {})
                        detected_arch = final.get('architecture', 'Unknown')
                        if final.get('bits') == 64:
                            detected_arch = f"{detected_arch}/64-bit"
                        if final.get('endian') == 'BE':
                            detected_arch = f"{detected_arch}/BE"
                else:
                    detected_arch = metadata.get("binary", "Unknown")
                    results["architecture_detection"] = {
                        "method_1": {"name": "Capstone Disassembly", "architecture": None, "confidence": 0},
                        "method_2": {"name": "Header + Prologue", "architecture": None, "confidence": 0},
                        "final": {"architecture": detected_arch, "confidence": 0, "method": "ghidra"}
                    }
                
                results["summary"] = {
                    "analysis_method": "ghidra_gnn",
                    "crypto_detected": ghidra_summary.get("crypto_detected", False),
                    "crypto_count": ghidra_summary.get("crypto_count", 0),
                    "protocols_detected": ghidra_summary.get("protocols_detected", 0),
                    "architecture": detected_arch,
                    "class_distribution": ghidra_summary.get("class_distribution", {})
                }
                
                # Include verifications if available
                if "verifications" in ghidra_result:
                    results["verifications"] = ghidra_result["verifications"]
                
                return results
                
        except Exception as e:
            import traceback
            print(f" Ghidra analysis failed: {e}")
            traceback.print_exc()
            results["ghidra_analysis"] = {"error": str(e)}
            # Fall back to heuristic analysis
    else:
        results["ghidra_analysis"] = {"available": False, "reason": "GHIDRA_PATH not set or Ghidra not found"}
    
    # ==================================================================
    # FALLBACK: Heuristic analysis (when Ghidra unavailable)
    # ==================================================================
    
    # Check if firmware needs extraction
    files_to_analyze = [file_path]
    extraction_performed = False
    extracted_files = []
    
    try:
        with open(file_path, 'rb') as f:
            header = f.read(1024)  # Read more for better detection
        
        # Calculate entropy to detect encrypted/compressed data
        file_entropy = calculate_entropy(header)
        
        # Detect if compressed/archived/firmware
        is_compressed = (
            header[:3] == b'\x5d\x00\x00' or  # LZMA
            header[:2] == b'\x1f\x8b' or       # gzip
            header[:4] == b'PK\x03\x04' or     # ZIP
            b'sqsh' in header or b'hsqs' in header or  # SquashFS
            header[:4] == b'HDR0' or           # Broadcom TRX
            header[:4] == b'\x27\x05\x19\x56' or  # uImage
            b'cramfs' in header or             # CramFS
            b'JFFS2' in header or              # JFFS2
            header[:2] == b'\x85\x19' or header[:2] == b'\x19\x85' or  # JFFS2 magic
            file_entropy > 7.5                 # High entropy = likely compressed/encrypted
        )
        
        # Also try extraction for common firmware extensions
        is_firmware = filename.lower().endswith(('.bin', '.img', '.fw', '.rom', '.trx', '.chk'))
        
        if is_compressed or is_firmware:
            extract_dir = UPLOAD_DIR / f"{job_id}_extracted"
            extract_dir.mkdir(exist_ok=True)
            
            # Try binwalk first
            binwalk_available = shutil.which('binwalk') is not None
            
            if binwalk_available:
                try:
                    result = subprocess.run(
                        ['binwalk', '-eM', '-C', str(extract_dir), file_path],
                        capture_output=True, timeout=120
                    )
                    extraction_performed = True
                    
                    # Find extracted binaries
                    for root, dirs, files in os.walk(extract_dir):
                        for f in files:
                            fpath = os.path.join(root, f)
                            fsize = os.path.getsize(fpath)
                            # Only include substantial files
                            if fsize > 1024 and not f.endswith(('.txt', '.log')):
                                extracted_files.append(fpath)
                except:
                    pass
            
            # Python-based extraction (no binwalk CLI needed)
            # Try multiple extraction methods
            
            # 1. LZMA
            if header[:3] == b'\x5d\x00\x00':
                try:
                    with open(file_path, 'rb') as f:
                        decompressed = lzma.decompress(f.read())
                    extracted_path = extract_dir / f"{filename}_lzma"
                    with open(extracted_path, 'wb') as f:
                        f.write(decompressed)
                    extracted_files.append(str(extracted_path))
                    extraction_performed = True
                except:
                    pass
            
            # 2. Gzip
            if header[:2] == b'\x1f\x8b':
                try:
                    with gzip.open(file_path, 'rb') as f:
                        decompressed = f.read()
                    extracted_path = extract_dir / f"{filename}_gzip"
                    with open(extracted_path, 'wb') as f:
                        f.write(decompressed)
                    extracted_files.append(str(extracted_path))
                    extraction_performed = True
                except:
                    pass
            
            # 3. ZIP
            if header[:4] == b'PK\x03\x04':
                try:
                    with zipfile.ZipFile(file_path, 'r') as z:
                        z.extractall(extract_dir)
                    for f in os.listdir(extract_dir):
                        extracted_files.append(str(extract_dir / f))
                    extraction_performed = True
                except:
                    pass
            
            # 4. Raw binary carving - scan for embedded content
            if not extracted_files:
                try:
                    with open(file_path, 'rb') as f:
                        full_data = f.read()
                    
                    # Look for embedded ELF files
                    elf_magic = b'\x7fELF'
                    offset = 0
                    elf_count = 0
                    while True:
                        idx = full_data.find(elf_magic, offset)
                        if idx == -1:
                            break
                        # Extract ELF (assume max 500KB per embedded file)
                        elf_data = full_data[idx:idx+512000]
                        extracted_path = extract_dir / f"embedded_elf_{elf_count}.elf"
                        with open(extracted_path, 'wb') as f:
                            f.write(elf_data)
                        extracted_files.append(str(extracted_path))
                        elf_count += 1
                        offset = idx + 1
                        if elf_count >= 5:
                            break
                    
                    # Look for embedded LZMA streams
                    lzma_magic = b'\x5d\x00\x00'
                    idx = full_data.find(lzma_magic)
                    if idx > 0 and idx < len(full_data) - 100:
                        try:
                            lzma_data = lzma.decompress(full_data[idx:])
                            extracted_path = extract_dir / f"embedded_lzma.bin"
                            with open(extracted_path, 'wb') as f:
                                f.write(lzma_data)
                            extracted_files.append(str(extracted_path))
                        except:
                            pass
                    
                    if extracted_files:
                        extraction_performed = True
                except:
                    pass
            
            if extracted_files:
                files_to_analyze = extracted_files[:5]  # Limit to 5 files
                results["extraction"] = {
                    "performed": True,
                    "method": "binwalk" if binwalk_available else "python",
                    "files_found": len(extracted_files),
                    "files_analyzed": len(files_to_analyze)
                }
    except Exception as e:
        results["extraction"] = {"error": str(e)}
    
    # Analyze all files (original or extracted)
    all_classifications = []
    arch_detected = "Unknown"
    arch_detection_details = None  # Store detailed architecture detection
    
    for fpath in files_to_analyze:
        try:
            crypto_detected, classifications = analyze_binary_heuristic(fpath)
            all_classifications.extend(classifications)
            
            # Try to detect architecture from extracted files using detailed method
            if arch_detected == "Unknown":
                arch_details = detect_architecture_detailed(fpath)
                if arch_details and arch_details.get('final', {}).get('architecture') != 'Unknown':
                    arch_detection_details = arch_details
                    final = arch_details.get('final', {})
                    arch = final.get('architecture', 'Unknown')
                    if final.get('bits') == 64:
                        arch = f"{arch}/64-bit"
                    if final.get('endian') == 'BE':
                        arch = f"{arch}/BE"
                    arch_detected = arch
        except:
            continue
    
    # Deduplicate classifications
    seen = set()
    unique_classifications = []
    for c in all_classifications:
        key = (c['name'], c['class_id'])
        if key not in seen:
            seen.add(key)
            unique_classifications.append(c)
    
    # ================================================================
    # ADVANCED ANALYSIS FOR ENCRYPTED/PROPRIETARY FIRMWARE
    # ================================================================
    firmware_intelligence = analyze_encrypted_firmware(file_path, unique_classifications)
    
    results["classifications"] = unique_classifications
    results["firmware_intelligence"] = firmware_intelligence
    
    # Add detailed architecture detection results
    if arch_detection_details:
        results["architecture_detection"] = {
            "method_1": {
                "name": arch_detection_details['method_1'].get('name', 'Capstone Disassembly'),
                "architecture": arch_detection_details['method_1'].get('architecture'),
                "confidence": arch_detection_details['method_1'].get('confidence', 0)
            },
            "method_2": {
                "name": arch_detection_details['method_2'].get('name', 'Header + Prologue'),
                "architecture": arch_detection_details['method_2'].get('architecture'),
                "confidence": arch_detection_details['method_2'].get('confidence', 0)
            },
            "final": arch_detection_details.get('final', {})
        }
    else:
        results["architecture_detection"] = {
            "method_1": {"name": "Capstone Disassembly", "architecture": None, "confidence": 0},
            "method_2": {"name": "Header + Prologue", "architecture": None, "confidence": 0},
            "final": {"architecture": arch_detected, "confidence": 0, "method": "fallback"}
        }
    
    results["summary"] = {
        "crypto_detected": any(c["class_id"] > 0 for c in unique_classifications) or firmware_intelligence.get("encryption_detected", False),
        "crypto_count": len([c for c in unique_classifications if c["class_id"] > 0]),
        "architecture": arch_detected,
        "files_analyzed": len(files_to_analyze),
        "extracted": extraction_performed,
        "encryption_status": firmware_intelligence.get("encryption_status", "unknown"),
        "security_level": firmware_intelligence.get("security_level", "unknown")
    }
    
    # Detect protocols
    results["protocols"] = detect_protocols(unique_classifications)
    
    return results


def analyze_encrypted_firmware(file_path, classifications):
    """
    Advanced analysis for encrypted/proprietary firmware.
    Provides useful intelligence even when extraction fails.
    """
    import struct
    
    try:
        with open(file_path, 'rb') as f:
            data = f.read()
    except:
        return {"error": "Could not read file"}
    
    analysis = {
        "file_size": len(data),
        "encryption_detected": False,
        "encryption_status": "unencrypted",
        "security_level": "low",
        "findings": [],
        "recommendations": []
    }
    
    # 1. Entropy analysis across file
    chunk_size = min(4096, len(data))
    entropies = []
    for i in range(0, len(data), chunk_size):
        chunk = data[i:i+chunk_size]
        ent = calculate_entropy(chunk)
        entropies.append(ent)
    
    avg_entropy = sum(entropies) / len(entropies) if entropies else 0
    max_entropy = max(entropies) if entropies else 0
    min_entropy = min(entropies) if entropies else 0
    entropy_variance = max_entropy - min_entropy
    
    analysis["entropy_analysis"] = {
        "average": round(avg_entropy, 3),
        "max": round(max_entropy, 3),
        "min": round(min_entropy, 3),
        "variance": round(entropy_variance, 3)
    }
    
    # 2. Determine encryption status
    if avg_entropy > 7.9:
        analysis["encryption_detected"] = True
        analysis["encryption_status"] = "fully_encrypted"
        analysis["security_level"] = "high"
        analysis["findings"].append({
            "type": "ENCRYPTION",
            "severity": "high",
            "description": f"Firmware is fully encrypted (entropy: {avg_entropy:.2f})",
            "confidence": 0.95
        })
    elif avg_entropy > 7.5:
        analysis["encryption_detected"] = True
        analysis["encryption_status"] = "compressed_or_encrypted"
        analysis["security_level"] = "medium"
        analysis["findings"].append({
            "type": "COMPRESSION",
            "severity": "medium",
            "description": f"Firmware uses strong compression or encryption (entropy: {avg_entropy:.2f})",
            "confidence": 0.85
        })
    elif entropy_variance > 3:
        analysis["encryption_status"] = "partially_encrypted"
        analysis["security_level"] = "medium"
        analysis["findings"].append({
            "type": "PARTIAL_ENCRYPTION",
            "severity": "medium",
            "description": "Firmware has mixed encrypted and unencrypted sections",
            "confidence": 0.75
        })
    
    # 3. Look for crypto signatures even in encrypted data
    crypto_indicators = [
        (b'AES', "AES encryption likely used"),
        (b'RSA', "RSA public key crypto likely used"),
        (b'SHA', "SHA hashing used"),
        (b'HMAC', "HMAC authentication used"),
        (b'CBC', "CBC block cipher mode"),
        (b'GCM', "GCM authenticated encryption"),
        (b'PKCS', "PKCS padding/format"),
        (b'X.509', "X.509 certificates"),
        (b'-----BEGIN', "PEM encoded data"),
        (b'ssh-rsa', "SSH authentication"),
        (b'ssl', "SSL/TLS protocol"),
        (b'tls', "TLS protocol"),
    ]
    
    for pattern, description in crypto_indicators:
        if pattern.lower() in data.lower():
            analysis["findings"].append({
                "type": "CRYPTO_INDICATOR",
                "severity": "info",
                "description": description,
                "confidence": 0.70
            })
    
    # 4. Firmware header analysis
    header = data[:256]
    if header[:4] == b'\x27\x05\x19\x56':
        analysis["findings"].append({
            "type": "FIRMWARE_TYPE",
            "severity": "info",
            "description": "uImage Linux kernel format detected",
            "confidence": 0.95
        })
    elif b'LZMA' in header or header[:3] == b'\x5d\x00\x00':
        analysis["findings"].append({
            "type": "COMPRESSION",
            "severity": "info",
            "description": "LZMA compression detected in header",
            "confidence": 0.90
        })
    elif b'sqsh' in header or b'hsqs' in header:
        analysis["findings"].append({
            "type": "FILESYSTEM",
            "severity": "info",
            "description": "SquashFS filesystem detected",
            "confidence": 0.95
        })
    
    # 5. Security recommendations based on findings
    if analysis["encryption_status"] == "fully_encrypted":
        analysis["recommendations"].append("Firmware uses strong encryption - indicates security-conscious design")
        analysis["recommendations"].append("Secure boot likely implemented")
        analysis["recommendations"].append("Cannot analyze without decryption key")
    elif analysis["encryption_status"] == "compressed_or_encrypted":
        analysis["recommendations"].append("Try firmware-specific unpacker for this vendor")
        analysis["recommendations"].append("Consider using JTAG/UART for analysis")
    else:
        analysis["recommendations"].append("Standard analysis should work")
    
    # 6. Overall assessment
    if len(analysis["findings"]) == 0 and avg_entropy > 7.5:
        analysis["findings"].append({
            "type": "PROPRIETARY",
            "severity": "medium",
            "description": "Proprietary or encrypted firmware format - limited visibility",
            "confidence": 0.80
        })
    
    return analysis


def analyze_binary_heuristic(file_path):
    """Analyze binary for crypto signatures using pattern matching."""
    with open(file_path, 'rb') as f:
        data = f.read()
    
    classifications = []
    crypto_found = False
    
    # ================================================================
    # STRING-BASED DETECTION - Works even in compressed firmware!
    # These strings often appear in firmware headers/metadata
    # ================================================================
    
    # Crypto library names (case-insensitive search)
    crypto_libs = [
        (b'openssl', "OpenSSL", 1, 0.90),
        (b'OpenSSL', "OpenSSL", 1, 0.92),
        (b'mbedtls', "mbedTLS", 1, 0.90),
        (b'mbedTLS', "mbedTLS", 1, 0.92),
        (b'wolfssl', "wolfSSL", 1, 0.90),
        (b'WolfSSL', "wolfSSL", 1, 0.92),
        (b'libsodium', "libsodium", 1, 0.88),
        (b'boringssl', "BoringSSL", 1, 0.90),
        (b'cryptlib', "CryptLib", 1, 0.85),
        (b'libtomcrypt', "LibTomCrypt", 1, 0.88),
        (b'nettle', "Nettle", 1, 0.85),
        (b'gnutls', "GnuTLS", 1, 0.88),
    ]
    
    for pattern, lib_name, class_id, conf in crypto_libs:
        if pattern in data:
            classifications.append({
                "name": f"{lib_name.lower()}_library",
                "class_id": class_id,
                "class_name": "AES/Block Cipher",
                "confidence": conf,
                "indicator": f"{lib_name} crypto library detected"
            })
            crypto_found = True
            break  # Only report one library
    
    # TLS/SSL string patterns
    tls_patterns = [
        (b'SSL_', "SSL Functions", 1, 0.85),
        (b'TLS_', "TLS Functions", 1, 0.85),
        (b'ssl_', "SSL Functions", 1, 0.82),
        (b'tls_', "TLS Functions", 1, 0.82),
        (b'HTTPS', "HTTPS Support", 1, 0.80),
        (b'https://', "HTTPS URLs", 1, 0.75),
        (b'certificate', "Certificate", 4, 0.78),
        (b'CERTIFICATE', "Certificate", 4, 0.80),
    ]
    
    for pattern, desc, class_id, conf in tls_patterns:
        if pattern in data:
            class_name = "AES/Block Cipher" if class_id == 1 else "Public Key"
            classifications.append({
                "name": f"ssl_marker_{desc.lower().replace(' ', '_')}",
                "class_id": class_id,
                "class_name": class_name,
                "confidence": conf,
                "indicator": f"{desc} found in firmware"
            })
            crypto_found = True
            break
    
    # Hash function names
    hash_patterns = [
        (b'SHA256', "SHA-256", 2, 0.90),
        (b'sha256', "SHA-256", 2, 0.88),
        (b'SHA-256', "SHA-256", 2, 0.92),
        (b'SHA512', "SHA-512", 2, 0.90),
        (b'SHA1', "SHA-1", 2, 0.85),
        (b'MD5', "MD5", 2, 0.80),
        (b'md5', "MD5", 2, 0.78),
        (b'HMAC', "HMAC", 5, 0.85),
        (b'hmac', "HMAC", 5, 0.82),
    ]
    
    for pattern, desc, class_id, conf in hash_patterns:
        if pattern in data:
            class_name = "Hash Function" if class_id == 2 else "Auth/MAC"
            classifications.append({
                "name": f"hash_{desc.lower().replace('-', '')}",
                "class_id": class_id,
                "class_name": class_name,
                "confidence": conf,
                "indicator": f"{desc} reference found"
            })
            crypto_found = True
    
    # Encryption algorithms
    enc_patterns = [
        (b'AES', "AES", 1, 0.85),
        (b'aes', "AES", 1, 0.80),
        (b'AES-128', "AES-128", 1, 0.90),
        (b'AES-256', "AES-256", 1, 0.92),
        (b'DES', "DES", 1, 0.75),
        (b'3DES', "3DES", 1, 0.78),
        (b'Blowfish', "Blowfish", 1, 0.82),
        (b'RC4', "RC4", 3, 0.80),
        (b'ChaCha', "ChaCha20", 3, 0.88),
        (b'chacha', "ChaCha20", 3, 0.85),
    ]
    
    for pattern, desc, class_id, conf in enc_patterns:
        if pattern in data:
            class_name = "AES/Block Cipher" if class_id == 1 else "Stream Cipher"
            classifications.append({
                "name": f"cipher_{desc.lower().replace('-', '')}",
                "class_id": class_id,
                "class_name": class_name,
                "confidence": conf,
                "indicator": f"{desc} algorithm reference"
            })
            crypto_found = True
    
    # ================================================================
    # BINARY PATTERN DETECTION - For uncompressed sections
    # ================================================================
    
    # Check for AES S-Box
    aes_sbox = bytes([0x63, 0x7c, 0x77, 0x7b, 0xf2, 0x6b, 0x6f, 0xc5])
    if aes_sbox in data:
        classifications.append({
            "name": "aes_sbox_detected",
            "class_id": 1,
            "class_name": "AES/Block Cipher",
            "confidence": 0.95,
            "indicator": "S-Box pattern found"
        })
        crypto_found = True
    
    # Check for SHA-256 constants
    sha256_k = bytes([0x42, 0x8a, 0x2f, 0x98])  # First K constant
    sha256_init = bytes([0x6a, 0x09, 0xe6, 0x67])  # H0
    if sha256_k in data or sha256_init in data:
        classifications.append({
            "name": "sha256_constants",
            "class_id": 2,
            "class_name": "Hash Function",
            "confidence": 0.92,
            "indicator": "SHA-256 constants found"
        })
        crypto_found = True
    
    # Check for MD5 constants
    md5_init = bytes([0x01, 0x23, 0x45, 0x67])
    if md5_init in data:
        classifications.append({
            "name": "md5_init",
            "class_id": 2,
            "class_name": "Hash Function",
            "confidence": 0.85,
            "indicator": "MD5 initialization vector"
        })
        crypto_found = True
    
    # Check for RSA/Public Key indicators
    rsa_patterns = [b'RSA', b'BEGIN PUBLIC KEY', b'BEGIN PRIVATE KEY', b'-----BEGIN']
    for pat in rsa_patterns:
        if pat in data:
            classifications.append({
                "name": "rsa_marker",
                "class_id": 4,
                "class_name": "Public Key",
                "confidence": 0.88,
                "indicator": f"Found: {pat.decode('utf-8', errors='ignore')}"
            })
            crypto_found = True
            break
    
    # Check for ChaCha/Salsa constants
    chacha_const = b'expand 32-byte k'
    if chacha_const in data:
        classifications.append({
            "name": "chacha_constant",
            "class_id": 3,
            "class_name": "Stream Cipher",
            "confidence": 0.94,
            "indicator": "ChaCha constant found"
        })
        crypto_found = True
    
    # Check for HMAC patterns
    if b'HMAC' in data or b'hmac' in data:
        classifications.append({
            "name": "hmac_marker",
            "class_id": 5,
            "class_name": "Auth/MAC",
            "confidence": 0.80,
            "indicator": "HMAC string found"
        })
        crypto_found = True
    
    # Check for random/entropy functions
    if b'rand' in data.lower() or b'random' in data.lower():
        classifications.append({
            "name": "prng_indicator",
            "class_id": 7,
            "class_name": "PRNG",
            "confidence": 0.70,
            "indicator": "Random function reference"
        })
        crypto_found = True
    
    # === LEGACY ROUTER CRYPTO (Linksys, D-Link, old firmware) ===
    
    # DES S-Box (legacy encryption)
    des_sbox = bytes([0x0e, 0x04, 0x0d, 0x01, 0x02, 0x0f, 0x0b, 0x08])
    if des_sbox in data:
        classifications.append({
            "name": "des_sbox",
            "class_id": 1,
            "class_name": "AES/Block Cipher",
            "confidence": 0.80,
            "indicator": "DES S-Box found (legacy)"
        })
        crypto_found = True
    
    # RC4 key schedule pattern (common in old routers)
    rc4_init = bytes(range(256))[:32]  # First 32 bytes of RC4 init
    if rc4_init in data:
        classifications.append({
            "name": "rc4_stream",
            "class_id": 3,
            "class_name": "Stream Cipher",
            "confidence": 0.85,
            "indicator": "RC4 key schedule detected"
        })
        crypto_found = True
    
    # CRC32 polynomial (used in many firmware)
    crc32_poly = bytes([0x04, 0xc1, 0x1d, 0xb7])  # CRC32 polynomial
    crc32_poly_le = bytes([0xb7, 0x1d, 0xc1, 0x04])
    if crc32_poly in data or crc32_poly_le in data:
        classifications.append({
            "name": "crc32_checksum",
            "class_id": 5,
            "class_name": "Auth/MAC",
            "confidence": 0.75,
            "indicator": "CRC32 polynomial found"
        })
        crypto_found = True
    
    # XOR Cipher detection (proprietary)
    xor_patterns = 0
    for i in range(0, min(len(data)-100, 10000), 100):
        chunk = data[i:i+100]
        xor_count = sum(1 for j in range(len(chunk)-1) if chunk[j] ^ chunk[j+1] in [0x00, 0xff, 0xaa, 0x55])
        if xor_count > 10:
            xor_patterns += 1
    if xor_patterns >= 3:
        classifications.append({
            "name": "xor_cipher",
            "class_id": 8,
            "class_name": "XOR Cipher",
            "confidence": 0.70,
            "indicator": "XOR encryption pattern detected"
        })
        crypto_found = True
    
    # Proprietary/Custom crypto (unknown patterns)
    # Look for repeating structures that suggest custom encryption
    entropy = calculate_entropy(data[:4096])
    if entropy > 7.5:  # High entropy = likely encrypted
        classifications.append({
            "name": "proprietary_encrypted",
            "class_id": 8,
            "class_name": "XOR Cipher",
            "confidence": 0.65,
            "indicator": f"High entropy ({entropy:.2f}) - possibly encrypted"
        })
        crypto_found = True
    
    # Blowfish P-array
    blowfish_p = bytes([0x24, 0x3f, 0x6a, 0x88])
    if blowfish_p in data:
        classifications.append({
            "name": "blowfish",
            "class_id": 1,
            "class_name": "AES/Block Cipher",
            "confidence": 0.82,
            "indicator": "Blowfish P-array found"
        })
        crypto_found = True
    
    # Kyber/Dilithium (Post-Quantum)
    kyber_q = bytes([0x01, 0x0d])  # Q = 3329 in LE
    if kyber_q in data:
        classifications.append({
            "name": "kyber_pqc",
            "class_id": 9,
            "class_name": "Post-Quantum",
            "confidence": 0.60,
            "indicator": "Kyber modulus detected"
        })
        crypto_found = True
    
    # If nothing found, mark as non-crypto
    if not classifications:
        classifications.append({
            "name": "no_crypto_detected",
            "class_id": 0,
            "class_name": "Non-Crypto",
            "confidence": 0.60,
            "indicator": "No known crypto patterns found"
        })
    
    return crypto_found, classifications


def calculate_entropy(data):
    """Calculate Shannon entropy of data."""
    import math
    if not data:
        return 0
    freq = {}
    for byte in data:
        freq[byte] = freq.get(byte, 0) + 1
    entropy = 0
    for count in freq.values():
        p = count / len(data)
        entropy -= p * math.log2(p)
    return entropy


def detect_architecture(file_path):
    """
    Detect binary architecture using multiple methods and return detailed results.
    
    Returns a dict with:
    - method_1: Capstone-based disassembly analysis (instruction validation)
    - method_2: Header + Prologue pattern matching
    - final: Ensemble consensus with highest confidence
    
    For backwards compatibility, if called as part of existing code expecting a string,
    convert using: result['final']['architecture'] or str(result)
    """
    try:
        result = detect_architecture_detailed(file_path)
        # Return just the final architecture string for backwards compatibility
        final = result.get('final', {})
        arch = final.get('architecture', 'Unknown')
        if final.get('bits') == 64:
            arch = f"{arch}/64-bit"
        if final.get('endian') == 'BE':
            arch = f"{arch}/BE"
        return arch
    except Exception:
        return "Unknown"


def detect_architecture_detailed(file_path):
    """
    Detect binary architecture using 2 independent methods and return detailed results.
    
    Returns:
        dict with 'method_1', 'method_2', and 'final' results
    """
    results = {
        'method_1': {'name': 'Capstone Disassembly', 'architecture': None, 'confidence': 0.0, 'details': {}},
        'method_2': {'name': 'Header + Prologue', 'architecture': None, 'confidence': 0.0, 'details': {}},
        'final': {'architecture': 'Unknown', 'confidence': 0.0, 'bits': 32, 'endian': 'LE', 'method': 'none'}
    }
    
    try:
        with open(file_path, 'rb') as f:
            data = f.read()
    except Exception as e:
        results['final']['error'] = str(e)
        return results
    
    # ================================================================
    # METHOD 1: Capstone-based strict disassembly analysis
    # ================================================================
    try:
        from arch_detection.capstone_detector import CapstoneDetector
        capstone_detector = CapstoneDetector()
        capstone_results = capstone_detector.detect(data)
        
        if capstone_results:
            best = capstone_results[0]
            results['method_1'] = {
                'name': 'Capstone Disassembly',
                'architecture': best.architecture,
                'confidence': round(best.confidence * 100, 1),
                'bits': best.bits,
                'endian': best.endian,
                'details': {
                    'coverage': best.details.get('valid_coverage', 0),
                    'continuity': best.details.get('continuity_score', 0),
                    'offset': best.offset
                }
            }
    except Exception as e:
        results['method_1']['error'] = str(e)
    
    # ================================================================
    # METHOD 2: Header parsing + Prologue pattern matching
    # ================================================================
    try:
        header_result = None
        prologue_result = None
        
        # Header detection (ELF/PE/Cortex-M)
        try:
            from arch_detection.header_detector import HeaderDetector
            header_detector = HeaderDetector()
            header_results = header_detector.detect(data)
            if header_results:
                header_result = header_results[0]
        except Exception:
            pass
        
        # Prologue pattern detection
        try:
            from arch_detection.prologue_detector import PrologueDetector
            prologue_detector = PrologueDetector()
            prologue_results = prologue_detector.detect(data)
            if prologue_results:
                prologue_result = prologue_results[0]
        except Exception:
            pass
        
        # Combine header + prologue
        if header_result and header_result.confidence > 0:
            results['method_2'] = {
                'name': 'Header + Prologue',
                'architecture': header_result.architecture,
                'confidence': round(header_result.confidence * 100, 1),
                'bits': header_result.bits,
                'endian': header_result.endian,
                'details': {
                    'source': header_result.details.get('source', 'header'),
                    'format': header_result.details.get('format', 'unknown')
                }
            }
        elif prologue_result and prologue_result.confidence > 0:
            results['method_2'] = {
                'name': 'Header + Prologue',
                'architecture': prologue_result.architecture,
                'confidence': round(prologue_result.confidence * 100, 1),
                'bits': prologue_result.bits,
                'endian': prologue_result.endian,
                'details': {
                    'source': 'prologue_patterns',
                    'matches': prologue_result.details.get('matches', 0)
                }
            }
        else:
            results['method_2']['details'] = {'source': 'none', 'reason': 'No headers or prologues found'}
    except Exception as e:
        results['method_2']['error'] = str(e)
    
    # ================================================================
    # FINAL: Ensemble consensus 
    # ================================================================
    try:
        from arch_detection.ensemble import detect_architecture_file
        
        ensemble_result = detect_architecture_file(file_path)
        if ensemble_result:
            results['final'] = {
                'architecture': ensemble_result.architecture,
                'confidence': round(ensemble_result.confidence * 100, 1),
                'bits': ensemble_result.bits,
                'endian': ensemble_result.endian,
                'method': 'ensemble',
                'agreement': {
                    'methods_agreed': ensemble_result.details.get('methods_agreed', 0),
                    'total_methods': ensemble_result.details.get('total_methods', 0),
                    'winning_detector': ensemble_result.details.get('winning_detector', 'unknown')
                }
            }
    except ImportError:
        # Fallback: Use best of method 1 or method 2
        m1_conf = results['method_1'].get('confidence', 0)
        m2_conf = results['method_2'].get('confidence', 0)
        
        if m1_conf >= m2_conf and m1_conf > 0:
            results['final'] = {
                'architecture': results['method_1']['architecture'],
                'confidence': m1_conf,
                'bits': results['method_1'].get('bits', 32),
                'endian': results['method_1'].get('endian', 'LE'),
                'method': 'capstone_fallback'
            }
        elif m2_conf > 0:
            results['final'] = {
                'architecture': results['method_2']['architecture'],
                'confidence': m2_conf,
                'bits': results['method_2'].get('bits', 32),
                'endian': results['method_2'].get('endian', 'LE'),
                'method': 'header_prologue_fallback'
            }
    except Exception as e:
        results['final']['error'] = str(e)
    
    # ================================================================
    # FIRMWARE-SPECIFIC FALLBACK (if ensemble didn't detect)
    # ================================================================
    if results['final'].get('architecture') in [None, 'Unknown'] or results['final'].get('confidence', 0) < 30:
        
        # ESP32/ESP8266 magic detection (Xtensa)
        if len(data) > 12 and data[0] == 0xE9:
            chip_id = data[12]
            chip_map = {
                0: ("ESP32", "Xtensa/ESP32-LX6"),
                2: ("ESP32-S2", "Xtensa/ESP32-S2-LX7"),
                5: ("ESP32-C3", "RISC-V/ESP32-C3"),
                9: ("ESP32-S3", "Xtensa/ESP32-S3-LX7"),
                12: ("ESP32-C2", "RISC-V/ESP32-C2"),
                13: ("ESP32-C6", "RISC-V/ESP32-C6"),
            }
            if chip_id in chip_map:
                chip_name, arch = chip_map[chip_id]
                results['final'] = {
                    'architecture': arch,
                    'confidence': 95.0,
                    'bits': 32,
                    'endian': 'LE',
                    'method': 'esp_magic'
                }
        
        # ELF format
        elif data[:4] == b'\x7fELF':
            arch_byte = data[18] if len(data) > 18 else 0
            arch_map = {
                3: "x86", 62: "x86_64", 40: "ARM", 183: "ARM64",
                8: "MIPS", 243: "RISC-V", 20: "PowerPC", 94: "Xtensa"
            }
            arch = arch_map.get(arch_byte, f"ELF-{arch_byte}")
            results['final'] = {
                'architecture': arch,
                'confidence': 98.0,
                'bits': 64 if arch_byte in [62, 183] else 32,
                'endian': 'BE' if data[5] == 2 else 'LE',
                'method': 'elf_header'
            }
        
        # PE/Windows
        elif data[:2] == b'MZ':
            results['final'] = {
                'architecture': "PE/x86",
                'confidence': 90.0,
                'bits': 32,
                'endian': 'LE',
                'method': 'pe_header'
            }
        
        # uImage (Linux)
        elif data[:4] == b'\x27\x05\x19\x56':
            if len(data) > 30:
                ih_arch = data[29]
                uimage_arch_map = {
                    2: "ARM", 3: "x86", 5: "MIPS", 6: "MIPS64",
                    7: "PowerPC", 15: "RISC-V", 22: "ARM64", 24: "Xtensa"
                }
                arch_name = uimage_arch_map.get(ih_arch, f"uImage-{ih_arch}")
                results['final'] = {
                    'architecture': f"uImage/{arch_name}",
                    'confidence': 95.0,
                    'bits': 64 if ih_arch in [6, 22] else 32,
                    'endian': 'BE',
                    'method': 'uimage_header'
                }
        
        # Intel HEX format (AVR/Arduino) - Capstone doesn't support AVR
        elif data[:1] == b':' and b':10' in data[:20]:
            results['final'] = {
                'architecture': "AVR/Intel-HEX",
                'confidence': 90.0,
                'bits': 8,
                'endian': 'LE',
                'method': 'intel_hex_format'
            }
        
        # Motorola S-Record (Z80, embedded) - Capstone doesn't support Z80
        elif data[:2] == b'S0' or data[:2] == b'S1':
            results['final'] = {
                'architecture': "Z80/S-Record",
                'confidence': 85.0,
                'bits': 8,
                'endian': 'LE',
                'method': 'srec_format'
            }
        
        # Z80 opcode detection - Capstone doesn't support Z80
        else:
            z80_score = 0
            if len(data) >= 64:
                if data[0:1] in [b'\xc3', b'\xc9', b'\x00']:  # JP, RET, NOP
                    z80_score += 1
                if b'\xc9' in data[:64]:  # RET instruction
                    z80_score += 1
                if b'\xc3' in data[:64]:  # JP instruction
                    z80_score += 1
                if b'\xcd' in data[:64]:  # CALL instruction
                    z80_score += 1
            
            if z80_score >= 3:
                results['final'] = {
                    'architecture': "Z80",
                    'confidence': 70.0,
                    'bits': 8,
                    'endian': 'LE',
                    'method': 'z80_opcodes'
                }
            
            # AVR detection (16-bit instructions, RJMP at start) - Capstone doesn't support AVR
            elif len(data) > 4 and data[0] >= 0x0c and data[0] <= 0x0f:
                if data[2] >= 0x0c and data[2] <= 0x0f:  # Multiple RJMP = vector table
                    results['final'] = {
                        'architecture': "AVR",
                        'confidence': 75.0,
                        'bits': 8,
                        'endian': 'LE',
                        'method': 'avr_rjmp_pattern'
                    }
            
            # Xtensa/ESP32 opcode detection - Capstone doesn't support Xtensa
            elif len(data) >= 256:
                xtensa_score = 0
                entry_count = 0
                call_count = 0
                retw_count = 0
                
                for i in range(min(256, len(data) - 2)):
                    byte0 = data[i]
                    byte1 = data[i + 1] if i + 1 < len(data) else 0
                    
                    # ENTRY instruction: op0=6, bits[7:4]=3 -> 0x36 pattern
                    if (byte0 & 0x0F) == 0x06 and (byte0 >> 4) == 0x03:
                        entry_count += 1
                    # CALL instructions: op0=5
                    if (byte0 & 0x0F) == 0x05:
                        call_count += 1
                    # RETW.N: 0x1D 0xF0
                    if byte0 == 0x1D and byte1 == 0xF0:
                        retw_count += 1
                
                if entry_count >= 3 or (entry_count >= 1 and call_count >= 5):
                    xtensa_score += 2
                if retw_count >= 2:
                    xtensa_score += 1
                
                if xtensa_score >= 2:
                    results['final'] = {
                        'architecture': "Xtensa/ESP32",
                        'confidence': 80.0,
                        'bits': 32,
                        'endian': 'LE',
                        'method': 'xtensa_opcodes'
                    }
        
        # ARM Cortex-M vector table detection (more reliable than Capstone for Cortex-M)
        if results['final'].get('architecture') in [None, 'Unknown'] and len(data) >= 8:
            import struct
            initial_sp = struct.unpack('<I', data[0:4])[0]
            reset_handler = struct.unpack('<I', data[4:8])[0]
            
            sp_in_sram = (initial_sp & 0xE0000000) == 0x20000000
            reset_is_thumb = (reset_handler & 0x1) == 1 and reset_handler < 0x20000000
            
            if sp_in_sram and reset_is_thumb:
                # Additional validation: Check for PUSH {LR} patterns
                push_lr_count = 0
                for i in range(0, min(256, len(data) - 1), 2):
                    if data[i+1] == 0xB5:  # PUSH {.., LR}
                        push_lr_count += 1
                
                if push_lr_count >= 2:
                    results['final'] = {
                        'architecture': "ARM/Cortex-M",
                        'confidence': 90.0,
                        'bits': 32,
                        'endian': 'LE',
                        'method': 'cortex_m_vector_table'
                    }
    
    return results


def detect_protocols(classifications):
    """Detect likely protocols based on crypto combinations."""
    protocols = []
    class_ids = {c["class_id"] for c in classifications}
    
    # TLS: AES + Hash + Public Key + MAC
    if 1 in class_ids and 2 in class_ids:
        protocols.append({
            "name": "TLS_HANDSHAKE",
            "description": "TLS/SSL Protocol detected",
            "confidence": 0.85 if 4 in class_ids else 0.70,
            "components": ["AES", "SHA256", "RSA" if 4 in class_ids else None]
        })
    
    # SSH: Similar to TLS
    if 1 in class_ids and 4 in class_ids:
        protocols.append({
            "name": "SSH_KEX",
            "description": "SSH Key Exchange",
            "confidence": 0.75,
            "components": ["Encryption", "PublicKey"]
        })
    
    # Secure Boot: Hash + Signature
    if 2 in class_ids and 4 in class_ids:
        protocols.append({
            "name": "SECURE_BOOT",
            "description": "Secure Boot Chain",
            "confidence": 0.70,
            "components": ["Hash", "Signature"]
        })
    
    return protocols


@app.get("/api/status/{job_id}")
async def get_status(job_id: str):
    result_path = RESULTS_DIR / f"{job_id}.json"
    if result_path.exists():
        return {"job_id": job_id, "status": "completed", "progress": 100}
    return {"job_id": job_id, "status": "pending", "progress": 0}


@app.get("/api/results/{job_id}")
async def get_results(job_id: str):
    result_path = RESULTS_DIR / f"{job_id}.json"
    if not result_path.exists():
        raise HTTPException(404, "Results not found")
    
    with open(result_path) as f:
        return json.load(f)


@app.get("/api/export/{job_id}/json")
async def export_json_endpoint(job_id: str):
    result_path = RESULTS_DIR / f"{job_id}.json"
    if not result_path.exists():
        raise HTTPException(404, "Results not found")
    return FileResponse(result_path, filename=f"cryptohunter_{job_id}.json")


@app.get("/api/export/{job_id}/csv")
async def export_csv_endpoint(job_id: str):
    """Export results as CSV file."""
    result_path = RESULTS_DIR / f"{job_id}.json"
    if not result_path.exists():
        raise HTTPException(404, "Results not found")
    
    import csv
    from io import StringIO
    
    with open(result_path) as f:
        data = json.load(f)
    
    output = StringIO()
    classifications = data.get("classifications", [])
    
    if classifications:
        writer = csv.DictWriter(output, fieldnames=["name", "class_id", "class_name", "confidence", "indicator"])
        writer.writeheader()
        for c in classifications:
            writer.writerow({
                "name": c.get("name", ""),
                "class_id": c.get("class_id", 0),
                "class_name": c.get("class_name", ""),
                "confidence": c.get("confidence", 0),
                "indicator": c.get("indicator", "")
            })
    
    csv_path = RESULTS_DIR / f"{job_id}.csv"
    with open(csv_path, 'w', newline='') as f:
        f.write(output.getvalue())
    
    return FileResponse(csv_path, filename=f"cryptohunter_{job_id}.csv", media_type="text/csv")


@app.get("/api/export/{job_id}/excel")
async def export_excel_endpoint(job_id: str):
    """Export results as Excel file (.xlsx)."""
    result_path = RESULTS_DIR / f"{job_id}.json"
    if not result_path.exists():
        raise HTTPException(404, "Results not found")
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV if openpyxl not available
        return await export_csv_endpoint(job_id)
    
    with open(result_path) as f:
        data = json.load(f)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Crypto Detection"
    
    # Header styling
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    # Headers
    headers = ["Function Name", "Class ID", "Class Name", "Confidence", "Indicator"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, c in enumerate(data.get("classifications", []), 2):
        ws.cell(row=row, column=1, value=c.get("name", ""))
        ws.cell(row=row, column=2, value=c.get("class_id", 0))
        ws.cell(row=row, column=3, value=c.get("class_name", ""))
        ws.cell(row=row, column=4, value=f"{c.get('confidence', 0)*100:.1f}%")
        ws.cell(row=row, column=5, value=c.get("indicator", ""))
    
    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    summary = data.get("summary", {})
    ws2.cell(row=1, column=1, value="Metric").font = Font(bold=True)
    ws2.cell(row=1, column=2, value="Value").font = Font(bold=True)
    ws2.cell(row=2, column=1, value="Crypto Detected")
    ws2.cell(row=2, column=2, value="Yes" if summary.get("crypto_detected") else "No")
    ws2.cell(row=3, column=1, value="Crypto Count")
    ws2.cell(row=3, column=2, value=summary.get("crypto_count", 0))
    ws2.cell(row=4, column=1, value="Architecture")
    ws2.cell(row=4, column=2, value=summary.get("architecture", "Unknown"))
    
    excel_path = RESULTS_DIR / f"{job_id}.xlsx"
    wb.save(excel_path)
    
    return FileResponse(excel_path, filename=f"cryptohunter_{job_id}.xlsx", 
                       media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.get("/api/export/{job_id}/pdf")
async def export_pdf_endpoint(job_id: str):
    """Export results as PDF report."""
    result_path = RESULTS_DIR / f"{job_id}.json"
    if not result_path.exists():
        raise HTTPException(404, "Results not found")
    
    with open(result_path) as f:
        data = json.load(f)
    
    # Generate HTML report (works without reportlab)
    summary = data.get("summary", {})
    classifications = data.get("classifications", [])
    protocols = data.get("protocols", [])
    
    html = f"""<!DOCTYPE html>
<html>
<head>
    <title>CryptoHunter Report - {job_id}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 40px; color: #333; }}
        h1 {{ color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }}
        h2 {{ color: #3498db; }}
        table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background: #3498db; color: white; }}
        tr:nth-child(even) {{ background: #f9f9f9; }}
        .summary-box {{ background: #ecf0f1; padding: 20px; border-radius: 8px; margin-bottom: 25px; }}
        .crypto {{ color: #27ae60; font-weight: bold; }}
        .non-crypto {{ color: #e74c3c; }}
    </style>
</head>
<body>
    <h1> CryptoHunter Analysis Report</h1>
    <p><strong>Job ID:</strong> {job_id}</p>
    <p><strong>Filename:</strong> {data.get('filename', 'N/A')}</p>
    <p><strong>Generated:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    
    <div class="summary-box">
        <h2> Summary</h2>
        <p><strong>Crypto Detected:</strong> <span class="{'crypto' if summary.get('crypto_detected') else 'non-crypto'}">{'Yes ' if summary.get('crypto_detected') else 'No'}</span></p>
        <p><strong>Total Crypto Primitives:</strong> {summary.get('crypto_count', 0)}</p>
        <p><strong>Architecture:</strong> {summary.get('architecture', 'Unknown')}</p>
        <p><strong>Security Level:</strong> {summary.get('security_level', 'Unknown').upper()}</p>
    </div>
    
    <h2> Detected Crypto Primitives</h2>
    <table>
        <tr><th>Name</th><th>Class</th><th>Confidence</th><th>Indicator</th></tr>
"""
    
    for c in classifications:
        html += f"<tr><td>{c.get('name', '')}</td><td>{c.get('class_name', '')}</td><td>{c.get('confidence', 0)*100:.1f}%</td><td>{c.get('indicator', '')}</td></tr>\n"
    
    html += "</table>"
    
    if protocols:
        html += "<h2> Detected Protocols</h2><ul>"
        for p in protocols:
            html += f"<li><strong>{p.get('name', '')}</strong>: {p.get('description', '')} ({p.get('confidence', 0)*100:.0f}%)</li>"
        html += "</ul>"
    
    html += "</body></html>"
    
    pdf_path = RESULTS_DIR / f"{job_id}_report.html"
    with open(pdf_path, 'w') as f:
        f.write(html)
    
    return FileResponse(pdf_path, filename=f"cryptohunter_{job_id}_report.html", 
                       media_type="text/html")


# ==============================================================
# WebSocket for Real-time Progress
# ==============================================================

@app.websocket("/ws/progress/{job_id}")
async def websocket_progress(websocket: WebSocket, job_id: str):
    """WebSocket endpoint for real-time analysis progress."""
    await manager.connect(websocket, job_id)
    try:
        while True:
            # Keep connection alive, wait for messages
            data = await websocket.receive_text()
            if data == "ping":
                await websocket.send_text("pong")
    except WebSocketDisconnect:
        manager.disconnect(websocket, job_id)


# ==============================================================
# Pipeline Steps Summary
# ==============================================================
# Step 0: Firmware Extraction (binwalk/unblob)
# Step 1: Ghidra P-Code Lifting
# Step 2: XGBoost Fast Filter (xgboost_filter.py) 
# Step 3: GNN Classification (infer_crypto.py)
# Step 4: Angr Symbolic Verification (symbolic_verify.py)
# Step 5: Protocol Detection & Aggregation


if __name__ == "__main__":
    import uvicorn
    print("\n" + "="*55)
    print("  CRYPTOHUNTER - AI/ML Crypto Detection Pipeline")
    print("="*55)
    print("  Pipeline Steps:")
    print("    0. Firmware Extraction (binwalk/unblob)")
    print("    1. Ghidra P-Code Lifting")
    print("    2. XGBoost Fast Filter")
    print("    3. GNN Classification")
    print("    4. Angr Symbolic Verification")
    print("    5. Protocol Detection")
    print("="*55)
    print(f"  Web UI: http://localhost:8000")
    print(f"  API Docs: http://localhost:8000/docs")
    print(f"  WebSocket: ws://localhost:8000/ws/progress/{{job_id}}")
    print("="*55 + "\n")
    uvicorn.run(app, host="0.0.0.0", port=8000)

